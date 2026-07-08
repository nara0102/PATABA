import os
from django.conf import settings
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
from django.template.loader import get_template
from xhtml2pdf import pisa
from django.http import HttpResponse, JsonResponse
from django.core.paginator import Paginator
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.utils import timezone
from django.db.models import Count, Sum, Q
from django.db import transaction
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib.auth.models import User
from pataba_core.utils import bersihkan_angka

# 1. pdf 
import qrcode
import base64
from io import BytesIO
from django.urls import reverse

# 2. Impor Model Milik Aset Tanah Sendiri
from .models import AsetTanah, FotoAsetTanah, RefKelurahan, RefKecamatan, SertifikatTanah, MasterOPD

# 3. Impor Hak Akses & Perekam Jejak dari Manajemen Pengguna
from apps.manajemen_pengguna.views import role_required, is_admin_bpkad, is_operator, catat_aktivitas
from apps.manajemen_pengguna.utils import get_user_profile

from pataba_core.constants import ROLE_OPERATOR, ROLE_ADMIN, STATUS_PENDING


    
# - - - - -
# CRUD ASET TANAH
# - - - - -

# 1. Input Aset Tanah Baru (SOP VALIDASI KETAT & PING-PONG)
@login_required(login_url='auth:login')
@role_required(allowed_roles=['ADMIN_BPKAD', 'OPERATOR_OPD'])
def input_aset_tanah(request):
    profile = request.user.profile
    role_user = profile.role.strip().upper()

    if request.method == 'POST':
        # a. TANGKAP DATA PENGIRIM & LOGIKA PING-PONG (Berdasarkan SOP SKILL.MD)
        if role_user == 'OPERATOR_OPD':
            target_opd = profile.opd
            target_status = 'BELUM_DIVERIFIKASI'
            catatan_revisi = ''  # SOP: Otomatis kosongkan catatan revisi saat OPD resubmit!
        else:
            opd_id = request.POST.get('opd')
            target_opd = MasterOPD.objects.get(pk=opd_id) if opd_id else None
            target_status = request.POST.get('status_verifikasi') or 'VALID'
            catatan_revisi = request.POST.get('catatan_revisi', '')

        # b. PEMBERSIH ANGKA RELASI & WILAYAH
        kec_id = request.POST.get('kecamatan')
        kel_id = request.POST.get('kelurahan')
        kec = RefKecamatan.objects.get(pk=kec_id) if kec_id else None
        kel = RefKelurahan.objects.get(pk=kel_id) if kel_id else None
        
        luas_bersih = bersihkan_angka(request.POST.get('luas_m2'), ke_integer=False) 
        
        nilai_clean = bersihkan_angka(request.POST.get('nilai_aset'), ke_integer=True)

        # c. PENYESUAIAN FIELD SERTIFIKAT (Menghindari Bentrok JS vs Python)
        status_sertifikasi = request.POST.get('status_sertifikasi') or 'BELUM_BERSERTIFIKAT'
        status_hak_sementara = request.POST.get('status_hak_sementara', '')
        keterangan_sertifikasi_lainnya = request.POST.get('keterangan_sertifikasi_lainnya', '')
        nama_pemegang_hak = request.POST.get('nama_pemegang_hak')
        
        # PENGAMAN LAPIS DUA (Backend): Memastikan data Pemkot tidak bocor jika JS gagal mengirim
        if status_sertifikasi == 'BERSERTIFIKAT':
            if 'PEMKOT' in keterangan_sertifikasi_lainnya:
                nama_pemegang_hak = 'Pemerintah Kota Palu'
        elif status_sertifikasi == 'BELUM_BERSERTIFIKAT':
            keterangan_sertifikasi_lainnya = request.POST.get('nomor_hak_sementara') or 'Dalam Proses Pembuatan'
            nama_pemegang_hak = '' # Kosongkan mutlak

        # d. INSTANCE LOKAL (Belum ter-save ke database)
        aset_baru = AsetTanah(
            opd=target_opd, 
            kelurahan=kel, 
            kecamatan_nama=kec.nama_kecamatan if kec else None, 
            kelurahan_nama=kel.nama_kelurahan if kel else None,
            nama_barang=request.POST.get('nama_barang'),
            kode_barang=request.POST.get('kode_barang'),
            nibar=request.POST.get('nibar'),
            nomor_register=request.POST.get('nomor_register'),
            status_kepemilikan=request.POST.get('status_kepemilikan') or 'BELUM_JELAS',
            luas_m2=luas_bersih,
            satuan='m2', 
            nilai_aset=nilai_clean,
            cara_perolehan=request.POST.get('cara_perolehan'),
            tanggal_perolehan=request.POST.get('tanggal_perolehan') or None,
            alamat_lokasi=request.POST.get('alamat_lokasi'),
            latitude=float(request.POST.get('latitude') or 0),
            longitude=float(request.POST.get('longitude') or 0),
            
            # --- Field Sertifikasi yang Sudah Disinkronkan ---
            status_sertifikasi=status_sertifikasi,
            nomor_sertifikat=request.POST.get('nomor_sertifikat'),
            status_hak_sementara=status_hak_sementara,
            nomor_hak=request.POST.get('nomor_hak'),
            keterangan_sertifikasi_lainnya=keterangan_sertifikasi_lainnya,
            nama_pemegang_hak=nama_pemegang_hak,
            
            kondisi_pemanfaatan=request.POST.get('kondisi_pemanfaatan'),
            status_penggunaan=request.POST.get('status_penggunaan'),
            spesifikasi_nama_barang=request.POST.get('spesifikasi_nama_barang'),
            spesifikasi_lainnya=request.POST.get('spesifikasi_lainnya'),
            keterangan=request.POST.get('keterangan'),
            
            status_verifikasi=target_status,
            catatan_revisi=catatan_revisi, 
            created_by_id=request.user.id
        )

        # e. VALIDASI ASET KETAT (Diperbarui)
        error_msgs = []
        if aset_baru.status_sertifikasi == 'BERSERTIFIKAT' and not aset_baru.nomor_sertifikat:
            error_msgs.append("Validasi Gagal: Nomor Sertifikat BPN wajib diisi karena status aset BERSERTIFIKAT.")
        elif aset_baru.status_sertifikasi == 'BELUM_BERSERTIFIKAT':
            if not aset_baru.status_hak_sementara:
                error_msgs.append("Validasi Gagal: Status Hak (Proses) wajib dipilih untuk aset Belum Bersertifikat (Pilih: Hak Pakai/Hak Milik).")
        
        # Validasi Khusus Logika Ping-Pong Admin BPKAD
        if role_user == 'ADMIN_BPKAD' and target_status == 'PERLU_REVIEW' and not catatan_revisi:
            error_msgs.append("Validasi Gagal: Karena Anda menolak usulan (Perlu Review), maka Catatan Revisi wajib diisi!")

        # Jika ada error, KEMBALIKAN KE TEMPLATE TANPA SAVE
        if error_msgs:
            for msg in error_msgs:
                messages.error(request, msg)
            return render(request, 'aset_tanah/input_aset_tanah.html', {
                'aset': aset_baru, 
                'opd_list': MasterOPD.objects.filter(is_active=1),
                'kecamatan_list': RefKecamatan.objects.all(),
                'kelurahan_list': RefKelurahan.objects.all(),
            })

        # f. SAVE KE DATABASE KALAU VALID
        try:
            with transaction.atomic(): # <--- KUNCI ATOMISITAS (Mencegah Data Hantu)
                aset_baru.save()
                
                # --- LOGIKA TANGKAP BANYAK FOTO ---
                daftar_foto = request.FILES.getlist('foto_aset_multiple') 
                if daftar_foto: # Pastikan foto ada sebelum di-looping
                    for file in daftar_foto:
                        FotoAsetTanah.objects.create(aset=aset_baru, file_foto=file)
            
            # (Pastikan fungsi catat_aktivitas sudah ada di luar fungsi ini)
            try:
                if role_user == 'OPERATOR_OPD':
                    catat_aktivitas(request.user, "Mengusulkan Aset Baru", aset_baru.nama_barang, request)
                else:
                    catat_aktivitas(request.user, "Menambahkan Master Aset", aset_baru.nama_barang, request)
            except Exception:
                pass # Abaikan jika catat_aktivitas belum terdefinisi sempurna
                
            if role_user == 'OPERATOR_OPD':
                messages.success(request, f"Usulan aset '{aset_baru.nama_barang}' berhasil dikirim dan menunggu verifikasi BPKAD.")
            else:
                if target_status == 'PERLU_REVIEW':
                    messages.warning(request, f"Usulan aset dikembalikan ke OPD untuk direvisi.")
                else:
                    messages.success(request, f"Aset '{aset_baru.nama_barang}' berhasil divalidasi dan ditambahkan ke master data publik.")
            return redirect('tanah:list_aset_tanah') # Pastikan namespace redirect sesuai
            
        except Exception as e:
            messages.error(request, f"Terjadi kesalahan database: {e}")
            return render(request, 'aset_tanah/input_aset_tanah.html', {
                'aset': aset_baru, 
                'opd_list': MasterOPD.objects.filter(is_active=1), 
                'kecamatan_list': RefKecamatan.objects.all(), 
                'kelurahan_list': RefKelurahan.objects.all()
            })

    # Saat GET (Membuka Form Kosong)
    return render(request, 'aset_tanah/input_aset_tanah.html', {
        'opd_list': MasterOPD.objects.filter(is_active=1),
        'kecamatan_list': RefKecamatan.objects.all(),
        'kelurahan_list': RefKelurahan.objects.all(),
    })

# 2. Daftar Aset Tanah
@login_required(login_url='auth:login')
@role_required(allowed_roles=['ADMIN_BPKAD', 'OPERATOR_OPD', 'SUPERADMIN'])
def list_aset_tanah(request):
    
    # role super admin
    if request.user.is_superuser:
        role_user = 'SUPERADMIN'
        base_template = 'base/base_superadmin.html'
    else:
        profile = request.user.profile
        role_user = profile.role.strip().upper()
        base_template = 'base/base_admin_BPKAD.html'
    
    # a. pembatasan hak akses
    if role_user == 'OPERATOR_OPD':
        queryset_dasar = AsetTanah.objects.filter(opd=request.user.profile.opd)
    else:
        # SUPERADMIN & ADMIN BPKAD bisa melihat semua data
        queryset_dasar = AsetTanah.objects.all()

    # b logika filter
    opd_id = request.GET.get('opd')
    kecamatan_id = request.GET.get('kecamatan')
    kelurahan_id = request.GET.get('kelurahan')
    status_verifikasi = request.GET.get('status_verifikasi')
    
    # Logika Filter Sertifikat Bertingkat
    status_sertifikasi = request.GET.get('status_sertifikasi')
    kriteria_sertifikat = request.GET.get('kriteria_sertifikat') # Untuk Asli/Fotokopi Pemkot/Non-Pemkot
    
    kondisi_pemanfaatan = request.GET.get('kondisi_pemanfaatan')
    status_kepemilikan = request.GET.get('status_kepemilikan')
    cara_perolehan = request.GET.get('cara_perolehan')
    koordinat = request.GET.get('koordinat')

    # Terapkan filter lanjutan ke 'queryset_dasar'
    if opd_id: queryset_dasar = queryset_dasar.filter(opd_id=opd_id)
    if kecamatan_id: queryset_dasar = queryset_dasar.filter(kelurahan__kecamatan_id=kecamatan_id)
    if kelurahan_id: queryset_dasar = queryset_dasar.filter(kelurahan_id=kelurahan_id)
    if status_verifikasi: queryset_dasar = queryset_dasar.filter(status_verifikasi=status_verifikasi)
    if status_sertifikasi: queryset_dasar = queryset_dasar.filter(status_sertifikasi=status_sertifikasi)
    
    # Menangkap sub-filter dari status "LAINNYA"
    if kriteria_sertifikat: 
        queryset_dasar = queryset_dasar.filter(keterangan_sertifikasi_lainnya=kriteria_sertifikat)
        
    if kondisi_pemanfaatan: queryset_dasar = queryset_dasar.filter(kondisi_pemanfaatan=kondisi_pemanfaatan)
    if status_kepemilikan: queryset_dasar = queryset_dasar.filter(status_kepemilikan=status_kepemilikan)
    if cara_perolehan: queryset_dasar = queryset_dasar.filter(cara_perolehan=cara_perolehan)
        
    if koordinat == 'ada':
        queryset_dasar = queryset_dasar.exclude(latitude=0).exclude(latitude__isnull=True)
    elif koordinat == 'belum':
        queryset_dasar = queryset_dasar.filter(Q(latitude=0) | Q(latitude__isnull=True))

   # statistik
    queryset_valid = queryset_dasar.filter(status_verifikasi='VALID')
    
    total = queryset_valid.count()
    agregasi = queryset_valid.aggregate(total_luas=Sum('luas_m2'), total_nilai=Sum('nilai_aset'))
    
    total_luas = agregasi['total_luas'] or 0
    total_nilai = agregasi['total_nilai'] or 0
    
    bersertifikat = queryset_valid.filter(status_sertifikasi='BERSERTIFIKAT').count()
    belum_sertifikat = queryset_valid.filter(status_sertifikasi='BELUM_BERSERTIFIKAT').count()
    persen_bersertifikat = (bersertifikat / total * 100) if total > 0 else 0
    
    ada_koordinat = queryset_valid.exclude(latitude=0).exclude(latitude__isnull=True).count()
    tanpa_koordinat = total - ada_koordinat
    
    # Antrean & Bermasalah tetap menghitung dari keseluruhan (bukan cuma yang valid)
    belum_verif = queryset_dasar.filter(status_verifikasi='BELUM_DIVERIFIKASI').count()
    bermasalah = queryset_dasar.filter(kondisi_pemanfaatan='SENGKETA').count()

    # pencarian kata kunci
    q = request.GET.get('q')
    semua_aset = queryset_dasar 
    
    if q:
        # Menghapus nama_lokasi_peruntukan yang sudah tiada agar tidak error
        semua_aset = semua_aset.filter(
            Q(opd__nama_opd__icontains=q) |
            Q(nama_barang__icontains=q) |
            Q(kode_barang__icontains=q) |
            Q(nibar__icontains=q) |
            Q(nomor_register__icontains=q) |
            Q(nomor_sertifikat__icontains=q) |
            Q(alamat_lokasi__icontains=q) |
            Q(kecamatan_nama__icontains=q) |
            Q(kelurahan_nama__icontains=q)
        )

    # Pengurutan data untuk tabel
    semua_aset = semua_aset.order_by('-created_at')
    
    # paginator
    paginator = Paginator(semua_aset, 10) # Angka 7 menentukan jumlah baris!
    page_number = request.GET.get('page')
    aset_page = paginator.get_page(page_number)

    # e kirim data ke template
    context = {
        'jenis_aset': 'Tanah',
        'semua_aset': aset_page, 
        'total_aset': total,
        'total_luas': total_luas,
        'total_nilai': total_nilai,
        'aset_bersertifikat': bersertifikat,
        'aset_belum_sertifikat': belum_sertifikat,
        'persen_bersertifikat': persen_bersertifikat,
        'aset_ada_koordinat': ada_koordinat,
        'aset_tanpa_koordinat': tanpa_koordinat,
        'aset_belum_verif': belum_verif,
        'aset_bermasalah': bermasalah,
        
        'opd_list': MasterOPD.objects.filter(is_active=1),
        'kecamatan_list': RefKecamatan.objects.all(),
        'kelurahan_list': RefKelurahan.objects.all(),
        
        'base_template': base_template,
    }
    
    return render(request, 'aset_tanah/list_aset_tanah.html', context)

# 3. Edit Aset Tanah
@login_required(login_url='auth:login')
@role_required(allowed_roles=['ADMIN_BPKAD', 'OPERATOR_OPD'])
def edit_aset_tanah(request, pk):
    aset = get_object_or_404(AsetTanah.objects.prefetch_related('koleksi_foto'), pk=pk)
    profile = request.user.profile
    role_user = profile.role.strip().upper()
    
    if request.user.is_superuser:
        messages.error(request, "Akses Ditolak: Akun Superadmin hanya memiliki hak akses Read-Only.")
        return redirect('tanah:list_aset_tanah')

    if role_user == 'OPERATOR_OPD' and aset.opd != profile.opd:
        messages.error(request, "Akses Ditolak: Anda tidak dapat mengubah data aset milik instansi lain.")
        return redirect('auth:dashboard_opd')

    if request.method == 'POST':
        # LOGIKA TIKET PING-PONG (STATUS OTOMATIS)
        if role_user == 'OPERATOR_OPD':
            target_opd = profile.opd
            target_status = 'BELUM_DIVERIFIKASI'
        else:
            opd_id = request.POST.get('opd')
            target_opd = MasterOPD.objects.get(pk=opd_id) if opd_id else aset.opd
            target_status = request.POST.get('status_verifikasi') or aset.status_verifikasi

        # UPDATE INSTANCE LOKAL (MENIMPA DATA SEMENTARA) ---
        kec_id = request.POST.get('kecamatan')
        kel_id = request.POST.get('kelurahan')
        kec = RefKecamatan.objects.get(pk=kec_id) if kec_id else aset.kelurahan.kecamatan
        kel = RefKelurahan.objects.get(pk=kel_id) if kel_id else aset.kelurahan
        
        aset.opd = target_opd
        aset.kelurahan = kel
        aset.kecamatan_nama = kec.nama_kecamatan
        aset.kelurahan_nama = kel.nama_kelurahan
        aset.nama_barang = request.POST.get('nama_barang')
        aset.kode_barang = request.POST.get('kode_barang')
        aset.nibar = request.POST.get('nibar')
        aset.nomor_register = request.POST.get('nomor_register')
        aset.status_kepemilikan = request.POST.get('status_kepemilikan') or 'BELUM_JELAS'
            
        luas_bersih = bersihkan_angka(request.POST.get('luas_m2'), ke_integer=False)
        aset.luas_m2 = luas_bersih
        nilai_clean = bersihkan_angka(request.POST.get('nilai_aset'), ke_integer=True)
        aset.nilai_aset = nilai_clean
        
        aset.cara_perolehan = request.POST.get('cara_perolehan')
        tanggal_perolehan = request.POST.get('tanggal_perolehan')
        if tanggal_perolehan: aset.tanggal_perolehan = tanggal_perolehan
        
        aset.alamat_lokasi = request.POST.get('alamat_lokasi')
        aset.latitude = float(request.POST.get('latitude') or 0)
        aset.longitude = float(request.POST.get('longitude') or 0)
        
        # PENYESUAIAN LOGIKA SERTIFIKAT (SINKRON DENGAN FUNGSI INPUT) 
        status_sertif = request.POST.get('status_sertifikasi') or 'BELUM_BERSERTIFIKAT'
        keterangan_sertifikasi_lainnya = request.POST.get('keterangan_sertifikasi_lainnya', '')
        nama_pemegang_hak = request.POST.get('nama_pemegang_hak')
        
        aset.status_sertifikasi = status_sertif
        aset.nomor_sertifikat = request.POST.get('nomor_sertifikat')
        aset.status_hak_sementara = request.POST.get('status_hak_sementara', '')
        aset.nomor_hak = request.POST.get('nomor_hak')
        
        # PENGAMAN LAPIS DUA (Backend) untuk proses Edit
        if status_sertif == 'BERSERTIFIKAT':
            aset.keterangan_sertifikasi_lainnya = keterangan_sertifikasi_lainnya
            if 'PEMKOT' in keterangan_sertifikasi_lainnya:
                aset.nama_pemegang_hak = 'Pemerintah Kota Palu'
            else:
                aset.nama_pemegang_hak = nama_pemegang_hak
        elif status_sertif == 'BELUM_BERSERTIFIKAT':
            aset.keterangan_sertifikasi_lainnya = request.POST.get('nomor_hak_sementara') or 'Dalam Proses Pembuatan'
            aset.nama_pemegang_hak = ''
        else:
            aset.keterangan_sertifikasi_lainnya = keterangan_sertifikasi_lainnya
            aset.nama_pemegang_hak = ''
        
        aset.kondisi_pemanfaatan = request.POST.get('kondisi_pemanfaatan')
        aset.status_penggunaan = request.POST.get('status_penggunaan')
        aset.spesifikasi_nama_barang = request.POST.get('spesifikasi_nama_barang')
        aset.spesifikasi_lainnya = request.POST.get('spesifikasi_lainnya')
        aset.keterangan = request.POST.get('keterangan')
        
        # LOGIKA CATATAN REVISI PING-PONG
        aset.status_verifikasi = target_status
        if role_user == 'ADMIN_BPKAD' or request.user.is_superuser:
            if target_status == 'PERLU_REVIEW':
                aset.catatan_revisi = request.POST.get('catatan_revisi')
            else:
                aset.catatan_revisi = "" 
        else:
            aset.catatan_revisi = "" 

        # GERBANG VALIDASI KEAMANAN & SOP PING-PONG 
        error_msgs = []
        if aset.status_sertifikasi == 'BERSERTIFIKAT' and not aset.nomor_sertifikat:
            error_msgs.append("Validasi Gagal: Nomor Sertifikat BPN wajib diisi karena status aset BERSERTIFIKAT.")
        elif aset.status_sertifikasi == 'BELUM_BERSERTIFIKAT':
            if not aset.status_hak_sementara:
                error_msgs.append("Validasi Gagal: Status Hak (Proses) wajib dipilih untuk aset Belum Bersertifikat (Pilih: Hak Pakai/Hak Milik).")

        if role_user == 'ADMIN_BPKAD' and target_status == 'PERLU_REVIEW':
            if not aset.catatan_revisi or aset.catatan_revisi.strip() == '':
                error_msgs.append("SOP BPKAD: Catatan revisi WAJIB diisi jika usulan dikembalikan ke OPD (PERLU REVIEW).")

        if error_msgs:
            for msg in error_msgs:
                messages.error(request, msg)
            return render(request, 'aset_tanah/input_aset_tanah.html', {
                'aset': aset, 
                'opd_list': MasterOPD.objects.filter(is_active=1),
                'kecamatan_list': RefKecamatan.objects.all(),
                'kelurahan_list': RefKelurahan.objects.all(),
                'daftar_foto': FotoAsetTanah.objects.filter(aset=aset), 
            })

        # EKSEKUSI SAVE DATABASE JIKA LOLOS VALIDASI 
        try:
            with transaction.atomic(): # <--- KUNCI ATOMISITAS
                aset.save()
                
                # bagian foto
                daftar_foto_baru = request.FILES.getlist('foto_aset_multiple') 
                if daftar_foto_baru:
                    for file in daftar_foto_baru:
                        FotoAsetTanah.objects.create(aset=aset, file_foto=file)
                        
                deleted_ids = request.POST.get('deleted_images')
                if deleted_ids:
                    list_id = deleted_ids.split(',')
                    # Ambil semua objek foto yang ID-nya ada di daftar hapus
                    fotos_to_delete = FotoAsetTanah.objects.filter(id__in=list_id, aset=aset)
                    
                    for foto in fotos_to_delete:
                        # Ini akan menghapus file fisik di Supabase (via Django Storage)
                        # dan menghapus record di database
                        foto.file_foto.delete(save=False) 
                        foto.delete()
            
            try:
                if role_user == 'ADMIN_BPKAD' or request.user.is_superuser:
                    if target_status == 'PERLU_REVIEW':
                        catat_aktivitas(request.user, "Menolak/Revisi Usulan Aset", aset.nama_barang, request)
                        messages.warning(request, f"Aset {aset.nama_barang} dikembalikan ke OPD untuk direvisi.")
                    elif target_status == 'VALID':
                        catat_aktivitas(request.user, "Memvalidasi Aset", aset.nama_barang, request)
                        messages.success(request, f"Aset {aset.nama_barang} berhasil diperbarui & diverifikasi.")
                    else:
                        catat_aktivitas(request.user, "Mengedit Data Aset", aset.nama_barang, request)
                        messages.success(request, f"Aset {aset.nama_barang} berhasil diperbarui.")
                else:
                    catat_aktivitas(request.user, "Mengirim Ulang Revisi Aset", aset.nama_barang, request)
                    messages.success(request, f"Usulan revisi aset {aset.nama_barang} berhasil dikirim ulang ke BPKAD.")
            except Exception:
                pass
                
            return redirect('tanah:list_aset_tanah') 
        except Exception as e:
            messages.error(request, f"Gagal memperbarui aset: {e}")
            return render(request, 'aset_tanah/input_aset_tanah.html', {
                'aset': aset, 
                'opd_list': MasterOPD.objects.filter(is_active=1), 
                'kecamatan_list': RefKecamatan.objects.all(),
                'kelurahan_list': RefKelurahan.objects.all(),
                'daftar_foto': FotoAsetTanah.objects.filter(aset=aset),
            })
    
    # Menampilkan Form Awal (Method GET)
    return render(request, 'aset_tanah/input_aset_tanah.html', {
        'aset': aset, 
        'opd_list': MasterOPD.objects.filter(is_active=1),
        'kecamatan_list': RefKecamatan.objects.all(),
        'kelurahan_list': RefKelurahan.objects.all(), # Ditambahkan agar kelurahan ter-load saat buka form
        'daftar_foto': FotoAsetTanah.objects.filter(aset=aset),
    })

# 4. Detail Aset Tanah
@login_required(login_url='auth:login')
@role_required(allowed_roles=['ADMIN_BPKAD', 'OPERATOR_OPD', 'SUPERADMIN'])
def detail_aset_tanah(request, pk):
    aset = get_object_or_404(AsetTanah.objects.prefetch_related('koleksi_foto'), pk=pk)
    
    if request.user.is_superuser:
        base_template = 'base/base_superadmin.html'
    else:
        base_template = 'base/base_admin_BPKAD.html'
        
    return render(request, 'aset_tanah/detail_aset_tanah.html', {
        'aset': aset,
        'profile': get_user_profile(request.user),
        'base_template': base_template
    })

# 5. Hapus Aset Tanah
@login_required(login_url='auth:login')
@role_required(allowed_roles=['ADMIN_BPKAD'])
def delete_aset_tanah(request, pk):
    aset = get_object_or_404(AsetTanah.objects.prefetch_related('koleksi_foto'), pk=pk)
    
    if request.user.is_superuser:
        messages.error(request, "Akses Ditolak: Akun Superadmin hanya memiliki hak akses Read-Only.")
        return redirect('tanah:list_aset_tanah')
    
    nama_aset_terhapus = aset.nama_barang
    aset.delete()
    catat_aktivitas(request.user, "Menghapus Data Aset", nama_aset_terhapus, request)
    messages.success(request, "Aset berhasil dihapus.")
    return redirect('tanah:list_aset_tanah')


# - - - - -
# CRUD SERTIFIKASI TANAH
# - - - - -

# 1. Tambah Sertifikat
@login_required(login_url='auth:login')
@role_required(allowed_roles=['ADMIN_BPKAD'])
def tambah_sertifikat(request, aset_id):
    if request.user.is_superuser:
        messages.error(request, "Akses Ditolak: Anda tidak memiliki otoritas mengubah dokumen legalitas BPN.")
        return redirect('tanah:list_sertifikat')
    
    aset = get_object_or_404(AsetTanah, pk=aset_id)
    
    if request.method == 'POST':
        luas_bersih = bersihkan_angka(request.POST.get('luas_m2'), ke_integer=False)
        nilai_clean = bersihkan_angka(request.POST.get('nilai_aset'), ke_integer=True) 

        opd_id = request.POST.get('opd')
        target_opd = MasterOPD.objects.get(pk=opd_id) if opd_id else None

        sertifikat_baru = SertifikatTanah(
            aset_tanah=aset,
            opd=target_opd,
            nomor_sertifikat=request.POST.get('nomor_sertifikat'),
            nomor_hak = request.POST.get('nomor_hak'),
            nama_pemegang_hak=request.POST.get('nama_pemegang_hak') or 'Pemerintah Kota Palu',
            alamat=request.POST.get('alamat'),
            peruntukan=request.POST.get('peruntukan'),
            status_hak=request.POST.get('status_hak'),
            luas=luas_bersih,
            tanggal_pembuatan = request.POST.get('tanggal_pembuatan') or None,
            tahun_terbit=request.POST.get('tahun_terbit') or None,
            nilai=nilai_clean,
            pemetaan_bpn=request.POST.get('pemetaan_bpn'),
            keterangan=request.POST.get('keterangan'),
            catatan=request.POST.get('catatan')
        )

        error_msgs = []
        if not sertifikat_baru.status_hak:
            error_msgs.append("Validasi Gagal: Status Hak Sertifikat wajib dipilih.")
        if not sertifikat_baru.nomor_sertifikat:
            error_msgs.append("Validasi Gagal: Nomor Sertifikat wajib diisi.")
        if not request.POST.get('nama_pemegang_hak'):
            error_msgs.append("Validasi Gagal: Nama Pemegang Hak wajib diisi (Sistem mendeteksi kekosongan).")
        if not sertifikat_baru.keterangan:
            error_msgs.append("Validasi Gagal: Bukti Fisik Sertifikat (Asli / Fotokopi) wajib dipilih.")
        if error_msgs:
            for msg in error_msgs:
                messages.error(request, msg)
            return render(request, 'aset_tanah/input_sertifikat.html', {
                'sertifikat': sertifikat_baru, 
                'aset': aset, 
                'opd_list': MasterOPD.objects.filter(is_active=1)
            })

        try:
            sertifikat_baru.save()
            
            # Otomatis ubah status aset master menjadi BERSERTIFIKAT
            aset.status_sertifikasi = 'BERSERTIFIKAT'
            aset.nomor_sertifikat = sertifikat_baru.nomor_sertifikat
            aset.nama_pemegang_hak = sertifikat_baru.nama_pemegang_hak
            aset.status_hak_sementara = sertifikat_baru.status_hak 
            aset.nomor_hak = sertifikat_baru.nomor_hak
            
            # Menyiapkan keterangan turunan untuk master aset
            kriteria_fisik = "ASLI" if sertifikat_baru.keterangan == 'ASLI' else "FOTOKOPI"
            kriteria_pemilik = "PEMKOT" if sertifikat_baru.nama_pemegang_hak == 'Pemerintah Kota Palu' else "NON_PEMKOT"
            aset.keterangan_sertifikasi_lainnya = f"{kriteria_fisik}_{kriteria_pemilik}"
            
            aset.save()
            catat_aktivitas(request.user, "Menerbitkan Sertifikat", f"{sertifikat_baru.nomor_sertifikat} ({aset.nama_barang})", request)
            messages.success(request, f"Sertifikat {sertifikat_baru.nomor_sertifikat} berhasil diterbitkan & disinkronkan ke Master Aset.")
            return redirect('tanah:detail_aset_tanah', pk=aset_id)
            
        except Exception as e:
            messages.error(request, f"Gagal menyimpan sertifikat: {e}")
            return render(request, 'aset_tanah/input_sertifikat.html', {'sertifikat': sertifikat_baru, 'aset': aset, 'opd_list': MasterOPD.objects.filter(is_active=1)})
        
    return render(request, 'aset_tanah/input_sertifikat.html', {'aset': aset, 'opd_list': MasterOPD.objects.filter(is_active=1)})

# 2. Edit Sertifikat
@login_required(login_url='auth:login')
@role_required(allowed_roles=['ADMIN_BPKAD'])
def edit_sertifikat(request, sertifikat_id):
    
    if request.user.is_superuser:
        messages.error(request, "Akses Ditolak: Anda tidak memiliki otoritas mengubah dokumen legalitas BPN.")
        return redirect('tanah:list_sertifikat')
    
    sertifikat = get_object_or_404(SertifikatTanah, pk=sertifikat_id)
    aset = sertifikat.aset_tanah
    
    if request.method == 'POST':
        luas_bersih = bersihkan_angka(request.POST.get('luas_m2'), ke_integer=False)
        sertifikat.luas = luas_bersih

        nilai_clean = bersihkan_angka(request.POST.get('nilai_aset'), ke_integer=True)
        sertifikat.nilai = nilai_clean

        opd_id = request.POST.get('opd')
        sertifikat.opd = MasterOPD.objects.get(pk=opd_id) if opd_id else None
        
        sertifikat.nomor_sertifikat = request.POST.get('nomor_sertifikat')
        sertifikat.nomor_hak = request.POST.get('nomor_hak')
        sertifikat.nama_pemegang_hak = request.POST.get('nama_pemegang_hak') or 'Pemerintah Kota Palu'
        sertifikat.alamat = request.POST.get('alamat')
        sertifikat.peruntukan = request.POST.get('peruntukan')
        sertifikat.status_hak = request.POST.get('status_hak')
        sertifikat.tanggal_pembuatan = request.POST.get('tanggal_pembuatan') or None
        sertifikat.tahun_terbit = request.POST.get('tahun_terbit') or None
        sertifikat.pemetaan_bpn = request.POST.get('pemetaan_bpn')
        sertifikat.keterangan = request.POST.get('keterangan')
        sertifikat.catatan = request.POST.get('catatan')
        
        error_msgs = []
        if not sertifikat.status_hak: error_msgs.append("Status Hak Sertifikat wajib dipilih.")
        if not sertifikat.nomor_sertifikat: error_msgs.append("Nomor Sertifikat wajib diisi.")
        if not request.POST.get('nama_pemegang_hak'): error_msgs.append("Nama Pemegang Hak wajib diisi.")
        if not sertifikat.keterangan: error_msgs.append("Bukti Fisik Sertifikat wajib dipilih.")

        if error_msgs:
            for msg in error_msgs:
                messages.error(request, msg)
            return render(request, 'aset_tanah/input_sertifikat.html', {'sertifikat': sertifikat, 'aset': aset, 'opd_list': MasterOPD.objects.filter(is_active=1)})
            
        try:
            sertifikat.save()
            
            # Sinkronisasi nomor sertifikat ke aset tanah jika berubah
            if aset.nomor_sertifikat != sertifikat.nomor_sertifikat or aset.nama_pemegang_hak != sertifikat.nama_pemegang_hak:
                aset.nomor_sertifikat = sertifikat.nomor_sertifikat
                aset.nama_pemegang_hak = sertifikat.nama_pemegang_hak
                aset.status_hak_sementara = sertifikat.status_hak
                aset.nomor_hak = sertifikat.nomor_hak
                
                # Update juga keterangan fisiknya
                kriteria_fisik = "ASLI" if sertifikat.keterangan == 'ASLI' else "FOTOKOPI"
                kriteria_pemilik = "PEMKOT" if sertifikat.nama_pemegang_hak == 'Pemerintah Kota Palu' else "NON_PEMKOT"
                aset.keterangan_sertifikasi_lainnya = f"{kriteria_fisik}_{kriteria_pemilik}"
                
                aset.save()

            catat_aktivitas(request.user, "Mengedit Data Sertifikat", f"{sertifikat.nomor_sertifikat} ({aset.nama_barang})", request)
            messages.success(request, "Data sertifikat berhasil diperbarui.")
            return redirect('tanah:detail_aset_tanah', pk=aset.id)
            
        except Exception as e:
            messages.error(request, f"Gagal memperbarui sertifikat: {e}")
            return render(request, 'aset_tanah/input_sertifikat.html', {'sertifikat': sertifikat, 'aset': aset, 'opd_list': MasterOPD.objects.filter(is_active=1)})
        
    return render(request, 'aset_tanah/input_sertifikat.html', {'sertifikat': sertifikat, 'aset': aset, 'opd_list': MasterOPD.objects.filter(is_active=1)})

# 3. Hapus Sertifikat
@login_required(login_url='auth:login')
@role_required(allowed_roles=['ADMIN_BPKAD'])
def delete_sertifikat(request, sertifikat_id):
    
    if request.user.is_superuser:
        messages.error(request, "Akses Ditolak: Anda tidak memiliki otoritas mengubah dokumen legalitas BPN.")
        return redirect('tanah:list_sertifikat')
    
    sertifikat = get_object_or_404(SertifikatTanah, pk=sertifikat_id)
    aset = sertifikat.aset_tanah
    nomor_sertifikat_terhapus = sertifikat.nomor_sertifikat
    nama_aset_terkait = aset.nama_barang
    
    sertifikat.delete()
    catat_aktivitas(request.user, "Menghapus Sertifikat", f"{nomor_sertifikat_terhapus} ({nama_aset_terkait})", request)
    messages.success(request, "Sertifikat berhasil dihapus.")

    return redirect('tanah:list_sertifikat')

# 4. List Sertifikat 
@login_required(login_url='auth:login')
@role_required(allowed_roles=['ADMIN_BPKAD', 'SUPERADMIN'])
def list_sertifikat(request):
    # --- LOGIKA AMAN UNTUK MEMBACA ROLE SUPERADMIN ---
    if request.user.is_superuser:
        role_user = 'SUPERADMIN'
        base_template = 'base/base_superadmin.html'
    else:
        profile = request.user.profile
        role_user = profile.role.strip().upper()
        base_template = 'base/base_admin_BPKAD.html'
    
    # a. pembatasan hak akses
    if role_user == 'OPERATOR_OPD':
        queryset_dasar = AsetTanah.objects.filter(opd=request.user.profile.opd)
    else:
        # SUPERADMIN & ADMIN BPKAD bisa melihat semua data
        queryset_dasar = AsetTanah.objects.all()
        
    semua_aset = AsetTanah.objects.filter(status_verifikasi='VALID')
    
    total_aset = semua_aset.count()
    luas_aset = semua_aset.aggregate(Sum('luas_m2'))['luas_m2__sum'] or 0
    nilai_aset = semua_aset.aggregate(Sum('nilai_aset'))['nilai_aset__sum'] or 0

    # Statistik Aset yang SUDAH Bersertifikat (Diambil dari yang VALID)
    aset_sertifikat = semua_aset.filter(status_sertifikasi='BERSERTIFIKAT')
    total_sertifikat = aset_sertifikat.count()
    luas_sertifikat = aset_sertifikat.aggregate(Sum('luas_m2'))['luas_m2__sum'] or 0
    nilai_sertifikat = aset_sertifikat.aggregate(Sum('nilai_aset'))['nilai_aset__sum'] or 0

    # Statistik Aset yang BELUM Bersertifikat / Lainnya
    aset_belum = semua_aset.exclude(status_sertifikasi='BERSERTIFIKAT')
    total_belum = aset_belum.count()
    luas_belum = aset_belum.aggregate(Sum('luas_m2'))['luas_m2__sum'] or 0
    nilai_belum = aset_belum.aggregate(Sum('nilai_aset'))['nilai_aset__sum'] or 0

    # Perhitungan Persentase Progres
    persen_sertifikat = (total_sertifikat / total_aset * 100) if total_aset > 0 else 0
    persen_belum = (total_belum / total_aset * 100) if total_aset > 0 else 0
    
    # Filter
    sertifikat_list = SertifikatTanah.objects.select_related('opd', 'aset_tanah').order_by('-created_at')
    
    q = request.GET.get('q')
    if q:
        sertifikat_list = sertifikat_list.filter(
            Q(nomor_sertifikat__icontains=q) |
            Q(aset_tanah__nama_barang__icontains=q) |
            Q(aset_tanah__nibar__icontains=q) |
            Q(opd__nama_opd__icontains=q)
        )
        
    keterangan = request.GET.get('keterangan')
    if keterangan:
        sertifikat_list = sertifikat_list.filter(keterangan=keterangan)
    pemegang_hak = request.GET.get('pemegang_hak')
    if pemegang_hak == 'PEMKOT':
        sertifikat_list = sertifikat_list.filter(nama_pemegang_hak__iexact='Pemerintah Kota Palu')
    elif pemegang_hak == 'LAINNYA':
        sertifikat_list = sertifikat_list.exclude(nama_pemegang_hak__iexact='Pemerintah Kota Palu')

    status_hak = request.GET.get('status_hak')
    if status_hak:
        sertifikat_list = sertifikat_list.filter(status_hak=status_hak)

    paginator = Paginator(sertifikat_list, 10) 
    page_number = request.GET.get('page')
    sertifikat_page = paginator.get_page(page_number)
    
    context = {
        'total_aset': total_aset,
        'luas_aset': luas_aset,
        'nilai_aset': nilai_aset,
        
        'total_sertifikat': total_sertifikat,
        'luas_sertifikat': luas_sertifikat,
        'nilai_sertifikat': nilai_sertifikat,
        
        'total_belum': total_belum,
        'luas_belum': luas_belum,
        'nilai_belum': nilai_belum,
        
        'persen_sertifikat': persen_sertifikat,
        'persen_belum': persen_belum,
    
        'semua_sertifikat': sertifikat_page,
        
        'base_template': base_template,
    }
    return render(request, 'aset_tanah/list_sertifikat.html', context)

# 5. Detail Sertifikat
@login_required(login_url='auth:login')
@role_required(allowed_roles=['ADMIN_BPKAD', 'SUPERADMIN'])
def detail_sertifikat(request, sertifikat_id):
    # Tarik data sertifikat beserta data aset tanah induknya
    sertifikat = get_object_or_404(SertifikatTanah, id=sertifikat_id)
    
    if request.user.is_superuser:
        base_template = 'base/base_superadmin.html'
    else:
        base_template = 'base/base_admin_BPKAD.html'
    
    return render(request, 'aset_tanah/detail_sertifikat.html', {
        'sertifikat': sertifikat,
        'aset': sertifikat.aset_tanah,
        'base_template': base_template
    })


# - - - - - -
# API Kec. Kel
# - - - - - -

# Dropdown Dinamis Kelurahan
@login_required(login_url='auth:login')
def get_kelurahan_by_kecamatan(request):
    kecamatan_id = request.GET.get('kecamatan_id')
    if kecamatan_id:
        kelurahan = RefKelurahan.objects.filter(kecamatan_id=kecamatan_id).values('id', 'nama_kelurahan')
        return JsonResponse(list(kelurahan), safe=False)
    return JsonResponse([], safe=False)


# aset kendaraan sementara
def list_aset_kendaraan(request):
    # Secara cerdas melempar pengguna ke rute coming_soon yang ada di portal_publik
    return redirect('publik:coming_soon_kendaraan')


# - - - - -
# PEMBEBASAN LAHAN (GANTI RUGI)
# - - - - -

@login_required(login_url='auth:login')
@role_required(allowed_roles=['ADMIN_BPKAD'])
def list_pembebasan_lahan(request):
    """
    Fungsi sementara (Mockup) untuk halaman Pembebasan Lahan.
    Data ditarik dari HTML statis untuk bahan diskusi dengan Dinas.
    """
    return render(request, 'aset_tanah/pembebasan_lahan.html')


# - - - - -
# Document
# - - - - -

# 1. Tampilkan Halaman Konfigurasi Export Excel 
@login_required(login_url='auth:login')
def halaman_export_excel(request):
    opd_list = MasterOPD.objects.filter(is_active=1)
    kecamatan_list = RefKecamatan.objects.all()
    
    return render(request, 'aset_tanah/reports/halaman_export_excel.html', {
        'opd_list': opd_list,
        'kecamatan_list': kecamatan_list,
    })

# 2. Mesin Pembuat Excel Dinamis (Berdasarkan Centang)

# fungsi ambil data dari sertifikat
def ambil_detail_sertifikat(aset, field_name, tipe_data='string'):
    sertifikat = aset.data_sertifikat.first()
    if sertifikat:
        nilai = getattr(sertifikat, field_name, None)
        if nilai is not None:
            if tipe_data == 'numeric':
                try:
                    return float(nilai)
                except (ValueError, TypeError):
                    return 0
            if tipe_data == 'date' and hasattr(nilai, 'strftime'):
                return nilai.strftime("%d-%m-%Y")
            return nilai
    return 0 if tipe_data == 'numeric' else "-"

# proses excel
@login_required(login_url='auth:login')
@role_required(allowed_roles=['ADMIN_BPKAD', 'SUPERADMIN'])
def proses_export_excel(request):
    if request.method == 'POST':
        q = request.POST.get('q', '')
        opd_id = request.POST.get('opd', '')
        status_sertif = request.POST.get('status_sertifikasi', '')
        status_verifikasi = request.POST.get('status_verifikasi', 'VALID') 
        bukti_fisik = request.POST.get('bukti_fisik', '')
        pemegang_hak = request.POST.get('pemegang_hak', '')

        kolom_pilihan = request.POST.getlist('kolom_export')
        
        # --- LOGIKA KONSOLIDASI PINTAR (ANTI-DUPLIKAT) ---
        
        def logika_gabungan_status_hak(a):
            if a.status_sertifikasi == 'BERSERTIFIKAT':
                sertif = a.data_sertifikat.first()
                if sertif and sertif.status_hak:
                    return "Hak Pakai" if sertif.status_hak == 'HAK_PAKAI' else "Hak Milik"
            return a.get_status_hak_sementara_display() or "-"

        def logika_gabungan_nomor_hak(a):
            if a.status_sertifikasi == 'BERSERTIFIKAT':
                sertif = a.data_sertifikat.first()
                return sertif.nomor_hak if (sertif and sertif.nomor_hak) else "-"
            return a.nomor_hak or "-"

        def logika_kondisi_fisik_dokumen(a):
            # Prioritas 1: Ambil dari data riil transaksi buku sertifikat anak
            sertif = a.data_sertifikat.first()
            if sertif and sertif.keterangan:
                return "Dokumen Asli" if sertif.keterangan == 'ASLI' else "Hanya Fotokopi"
                
            # Prioritas 2: Ambil dari kode sinkronisasi master 5 pilar (Toleran semua format tulisan)
            if a.keterangan_sertifikasi_lainnya:
                kl_upper = a.keterangan_sertifikasi_lainnya.upper()
                if 'ASLI' in kl_upper: 
                    return 'Dokumen Asli'
                if 'FOTOKOPI' in kl_upper or 'FOTO_KOPI' in kl_upper or 'FC' in kl_upper: 
                    return 'Hanya Fotokopi'
            return "-"

        # --- PETA KOLOM BERSIH (FIXED) ---
        PETA_KOLOM = {
            'opd': ('Instansi (OPD)', lambda a: a.opd.nama_opd if a.opd else "-", None),
            'kode_barang': ('Kode Barang', lambda a: a.kode_barang or "-", None),
            'nama_barang': ('Nama Barang', lambda a: a.nama_barang or "-", None),
            'nibar': ('NIBAR', lambda a: a.nibar or "-", None),
            'nomor_register': ('Nomor Register', lambda a: a.nomor_register or "-", None),
            'luas': ('Luas Tanah (m2)', lambda a: float(a.luas_m2) if a.luas_m2 else 0, '#,##0 "m²"'),
            'nilai': ('Nilai Aset (Rp)', lambda a: float(a.nilai_aset) if a.nilai_aset else 0, '"Rp"#,##0'),
            'alamat': ('Alamat Lokasi', lambda a: a.alamat_lokasi or "-", None),
            'kecamatan': ('Kecamatan', lambda a: a.kecamatan_nama or "-", None),
            'kelurahan': ('Kelurahan', lambda a: a.kelurahan_nama or "-", None),
            'koordinat': ('Titik Koordinat', lambda a: f"{a.latitude}, {a.longitude}" if a.latitude and a.longitude else "-", None),
            'cara_perolehan': ('Cara Perolehan', lambda a: a.cara_perolehan or "-", None),
            'tgl_perolehan': ('Tanggal Perolehan', lambda a: a.tanggal_perolehan.strftime("%d-%m-%Y") if a.tanggal_perolehan else "-", None),
            'status_guna': ('Status Penggunaan', lambda a: a.status_penggunaan or "-", None),
            'kondisi': ('Kondisi Pemanfaatan', lambda a: a.kondisi_pemanfaatan or "-", None),
            
            # Kolom Hasil Penyatuan (Deduplicated)
            'status_sertif': ('Status Sertifikasi', lambda a: a.get_status_sertifikasi_display() or "-", None),
            'status_hak': ('Status Hak (HP/HM)', logika_gabungan_status_hak, None),
            'nomor_hak': ('Nomor Hak', logika_gabungan_nomor_hak, None),
            'kondisi_fisik': ('Kondisi Fisik Dokumen', logika_kondisi_fisik_dokumen, None),
            
            # Sisa Rincian Tambahan Buku Sertifikat (BPN)
            'sertif_nomor': ('Nomor Sertifikat (BPN)', lambda a: ambil_detail_sertifikat(a, 'nomor_sertifikat'), None),
            'sertif_pemegang': ('Nama Pemegang Hak', lambda a: ambil_detail_sertifikat(a, 'nama_pemegang_hak'), None),
            'sertif_alamat': ('Alamat di Sertifikat', lambda a: ambil_detail_sertifikat(a, 'alamat'), None),
            'sertif_peruntukan': ('Peruntukan Wilayah BPN', lambda a: ambil_detail_sertifikat(a, 'peruntukan'), None),
            'sertif_luas': ('Luas Sertifikat (BPN)', lambda a: ambil_detail_sertifikat(a, 'luas', 'numeric'), '#,##0 "m²"'),
            'sertif_nilai': ('Nilai Sertifikat (BPN)', lambda a: ambil_detail_sertifikat(a, 'nilai', 'numeric'), '"Rp"#,##0'),
            'sertif_tgl_buat': ('Tanggal Pembuatan Sertifikat', lambda a: ambil_detail_sertifikat(a, 'tanggal_pembuatan', 'date'), None),
            'sertif_tahun': ('Tahun Terbit Sertifikat', lambda a: ambil_detail_sertifikat(a, 'tahun_terbit'), None),
            'sertif_pemetaan': ('Pemetaan BPN', lambda a: ambil_detail_sertifikat(a, 'pemetaan_bpn'), None),
            'sertif_catatan': ('Catatan Tambahan', lambda a: ambil_detail_sertifikat(a, 'catatan'), None),
        }

        if not kolom_pilihan:
            kolom_pilihan = ['opd', 'kode_barang', 'nama_barang', 'nilai']

        headers = ['No'] + [PETA_KOLOM[k][0] for k in kolom_pilihan if k in PETA_KOLOM]

        aset_data = AsetTanah.objects.all().select_related('opd').prefetch_related('data_sertifikat')
        
        # Jalankan filter query
        if status_verifikasi != 'SEMUA':
            aset_data = aset_data.filter(status_verifikasi=status_verifikasi)
        if q:
            aset_data = aset_data.filter(Q(nama_barang__icontains=q) | Q(nibar__icontains=q) | Q(alamat_lokasi__icontains=q))
        if opd_id:
            aset_data = aset_data.filter(opd_id=opd_id)
        if status_sertif:
            aset_data = aset_data.filter(status_sertifikasi=status_sertif)
            
        if bukti_fisik == 'ASLI':
            aset_data = aset_data.filter(keterangan_sertifikasi_lainnya__startswith='ASLI')
        elif bukti_fisik == 'FOTOKOPI':
            aset_data = aset_data.filter(keterangan_sertifikasi_lainnya__startswith='FOTOKOPI')
            
        if pemegang_hak == 'PEMKOT':
            aset_data = aset_data.filter(keterangan_sertifikasi_lainnya__endswith='PEMKOT')
        elif pemegang_hak == 'LAINNYA':
            aset_data = aset_data.filter(keterangan_sertifikasi_lainnya__endswith='NON_PEMKOT')

        aset_data = aset_data.order_by('-created_at')

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="Custom_Laporan_Aset_BPN_PATABA.xlsx"'
        
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = 'Data Inventaris'

        header_fill = PatternFill(start_color="002D5E", end_color="002D5E", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        
        for col_num, header_title in enumerate(headers, 1):
            cell = worksheet.cell(row=1, column=col_num, value=header_title)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')

        for row_num, aset in enumerate(aset_data, 2):
            row_data = [row_num - 1]
            formats = [None]
            
            for k in kolom_pilihan:
                if k in PETA_KOLOM:
                    row_data.append(PETA_KOLOM[k][1](aset))
                    formats.append(PETA_KOLOM[k][2])

            for col_num, cell_value in enumerate(row_data, 1):
                cell = worksheet.cell(row=row_num, column=col_num, value=cell_value)
                fmt = formats[col_num - 1]
                if fmt:
                    cell.number_format = fmt
                    cell.alignment = Alignment(horizontal='right')

        for col in worksheet.columns:
            max_length = max(len(str(cell.value or '')) for cell in col)
            col_letter = col[0].column_letter
            worksheet.column_dimensions[col_letter].width = max(max_length + 3, 10)

        workbook.save(response)
        return response

    return redirect('tanah:halaman_export_excel')

# 3. Halaman Import Excel
@login_required(login_url='auth:login')
@role_required(allowed_roles=['ADMIN_BPKAD', 'SUPERADMIN'])
def halaman_import_view(request):
    # Hanya bertugas membuka halaman utama import
    return render(request, 'aset_tanah/reports/halaman_import.html')

# 4. Template file import excel
@login_required(login_url='auth:login')
@role_required(allowed_roles=['ADMIN_BPKAD', 'SUPERADMIN'])
def unduh_template_import(request):
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = 'Template_Import_Aset'

    # 1. HEADER DIBERI PERINGATAN AGAR MELIHAT SHEET 2
    headers = [
        'ID OPD (Angka)', 'Nama Barang', 'Kode Barang', 'NIBAR', 'Nomor Register', 
        'Status Kepemilikan (Cek Sheet 2)', 'Luas (tanpa titik dan m2)', 'Cara Perolehan (Cek Sheet 2)', 'Tanggal Perolehan (DD/MM/YYYY)', 
        'Nilai Aset (Rp)', 'Latitude', 'Longitude', 'Nama Kecamatan', 'Nama Kelurahan', 
        'Alamat Lokasi', 'Status Sertifikasi Utama (Cek Sheet 2)', 'Nomor Sertifikat', 'Kondisi Pemanfaatan (Cek Sheet 2)', 
        'Status Penggunaan', 'Status Verifikasi (Isi: VALID)', 'Spesifikasi Nama Barang', 'Spesifikasi Lainnya',
        'Status Hak Sementara (Cek Sheet 2)', 'Nomor Hak', 'Nama Pemegang Hak', 'Keterangan Sertifikasi Khusus (Cek Sheet 2)'
    ]

    dummy_data = [
        '18', 'Tanah Bangunan Kantor Pemerintah', '1.3.1.01.01.04.001', '7271199550201001001131010104001000001', '0001',
        'PEMDA', '1563', 'HIBAH', '31/12/1995', 
        2910306000, '-0.8917', '119.8707', 'Mantikulore', 'Tanamodindi',
        'Jl. Baruga', 'BERSERTIFIKAT', '19.05.08.06.4.00076 AAH 951892', 'DIMANFAATKAN',
        'Badan Pengelola Keuangan Dan Aset Daerah 5.02.01.001.001', 'VALID', 'Tanah Bangunan Kantor Pemerintah', 'Kantor Badan Pengelola Keuangan dan Aset Daerah',
        'HAK_PAKAI', 'No. 00076', 'Pemerintah Kota Palu', 'ASLI_PEMKOT'
    ]

    # Styling Header Sheet 1
    header_fill = PatternFill(start_color="002D5E", end_color="002D5E", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    
    for col_num, header_title in enumerate(headers, 1):
        cell = worksheet.cell(row=1, column=col_num, value=header_title)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        
    for col_num, cell_value in enumerate(dummy_data, 1):
        cell = worksheet.cell(row=2, column=col_num, value=cell_value)
        cell.font = Font(color="808080", italic=True) 
        
    for row in range(1, 2005): # Mengunci hingga 2000 baris ke bawah
        worksheet.cell(row=row, column=3).number_format = '@' # Kode Barang
        worksheet.cell(row=row, column=4).number_format = '@' # NIBAR
        worksheet.cell(row=row, column=5).number_format = '@' # Nomor Register
        worksheet.cell(row=row, column=10).number_format = '"Rp"#,##0'# Nilai Aset
    
    # ukuran kolom   
    for col in worksheet.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col[:2]: 
            if cell.value:
                if cell.column == 10:
                    max_len = max(max_len, len(str(cell.value)) + 6)
                else:
                    max_len = max(max_len, len(str(cell.value)))
        worksheet.column_dimensions[col_letter].width = max(max_len + 5, 15)

    # ==============================================================
    # 2. SHEET KEDUA: KAMUS REFERENSI LENGKAP
    # ==============================================================
    ws_ref = workbook.create_sheet(title="Kamus Referensi Pilihan")
    
    # --- BAGIAN A: KAMUS OPD ---
    ws_ref.cell(row=1, column=1, value="DAFTAR ID INSTANSI (OPD)").font = Font(bold=True, color="002D5E")
    ws_ref.cell(row=2, column=1, value="ID OPD").font = Font(bold=True)
    ws_ref.cell(row=2, column=2, value="NAMA INSTANSI").font = Font(bold=True)
    
    opd_list = MasterOPD.objects.filter(is_active=1).order_by('nama_opd')
    for index, opd in enumerate(opd_list, start=3):
        ws_ref.cell(row=index, column=1, value=opd.id).alignment = Alignment(horizontal='center')
        ws_ref.cell(row=index, column=2, value=opd.nama_opd)
        
    ws_ref.column_dimensions['A'].width = 15
    ws_ref.column_dimensions['B'].width = 50

    # --- BAGIAN B: KAMUS PILIHAN CHOICES ---
    ws_ref.cell(row=1, column=4, value="DAFTAR PILIHAN BAKU (pastikan mengisi data sesuai dengan format pilihan)").font = Font(bold=True, color="002D5E")
    
    # Dictionary berisi semua pilihan baku sesuai models.py kita
    kamus_pilihan = {
        'Status Kepemilikan (Kolom F)': ['BELUM_JELAS', 'PEMDA', 'MASYARAKAT', 'PROVINSI', 'PUSAT'],
        'Cara Perolehan (Kolom H)': ['HIBAH', 'PEMBELIAN', 'PENYERAHAN', 'GANTI_RUGI'],
        'Status Sertifikasi (Kolom P)': ['BERSERTIFIKAT', 'BELUM_BERSERTIFIKAT'],
        'Kondisi Pemanfaatan (Kolom R)': ['TIDAK_DIKETAHUI', 'DIMANFAATKAN', 'TANAH_KOSONG', 'RUSAK', 'SENGKETA'],
        'Status Hak Sementara (Kolom W)': ['HAK_PAKAI', 'HAK_MILIK'],
        'Keterangan Sertif (Kolom Z)': ['ASLI_PEMKOT', 'FOTOKOPI_PEMKOT', 'ASLI_NON_PEMKOT', 'FOTOKOPI_NON_PEMKOT']
    }

    start_col = 4
    for header, pilihan in kamus_pilihan.items():
        # Buat Header Pilihan
        cell_header = ws_ref.cell(row=2, column=start_col, value=header)
        cell_header.font = Font(bold=True)
        cell_header.fill = PatternFill(start_color="E2E8F0", end_color="E2E8F0", fill_type="solid")
        
        # Cetak isi pilihannya ke bawah
        for r_idx, opsi in enumerate(pilihan, start=3):
            ws_ref.cell(row=r_idx, column=start_col, value=opsi)
            
        ws_ref.column_dimensions[ws_ref.cell(row=2, column=start_col).column_letter].width = 28
        start_col += 1
    # ==============================================================

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="Template Import Data Aset Tanah.xlsx"'
    workbook.save(response)
    
    return response

# 5. Mesin Proses Import Excel
@login_required(login_url='auth:login')
@role_required(allowed_roles=['ADMIN_BPKAD', 'SUPERADMIN'])
def proses_import_excel(request):
    if request.method == 'POST' and request.FILES.get('file_excel'):
        file_excel = request.FILES['file_excel']
        
        # Validasi ekstensi file keamanan
        if not file_excel.name.endswith('.xlsx'):
            messages.error(request, "Validasi Gagal: Format file wajib berupa .xlsx (Excel Modern).")
            return redirect('tanah:halaman_import')

        try:
            # Membaca berkas langsung dari memori request browser
            wb = openpyxl.load_workbook(file_excel, data_only=True)
            sheet = wb.active
            
            berhasil = 0
            gagal = 0
            baris_error_logs = []

            # Gunakan transaksi atomis: jika di tengah jalan ada baris rusak, batalkan semua!
            with transaction.atomic():
                for i, row in enumerate(sheet.iter_rows(min_row=3, values_only=True), start=3):
                    if not row or not row[0]: 
                        break

                    opd_id = row[0]
                    nama_barang = row[1]
                    kode_barang = row[2]
                    nibar = row[3]
                    
                    # Logika penataan nomor register zfill(4)
                    nomor_register_raw = row[4]
                    if nomor_register_raw is not None:
                        if isinstance(nomor_register_raw, (int, float)):
                            nomor_register = str(int(nomor_register_raw)).strip().zfill(4)
                        else:
                            nomor_register = str(nomor_register_raw).strip().zfill(4)
                    else:
                        nomor_register = "0000"

                    status_kepemilikan = row[5]
                    luas_m2 = bersihkan_angka(row[6], ke_integer=False)
                    cara_perolehan = row[7]
                    tanggal_perolehan = row[8] 
                    nilai_aset = bersihkan_angka(row[9], ke_integer=True)
                    latitude = row[10]
                    longitude = row[11]
                    kecamatan_nama = row[12]
                    kelurahan_nama = row[13]
                    alamat_lokasi = row[14]
                    
                    # Pencarian objek kelurahan berbasis teks database
                    kel_obj = None
                    if kelurahan_nama:
                        kel_obj = RefKelurahan.objects.filter(nama_kelurahan__iexact=kelurahan_nama).first()
                    
                    status_sertifikasi = row[15] if row[15] else 'BELUM_BERSERTIFIKAT'
                    nomor_sertifikat = row[16]
                    kondisi_pemanfaatan = row[17]
                    status_penggunaan = row[18]
                    status_verifikasi = row[19] if row[19] else 'VALID'

                    # Aturan Lapis Batas Panjang Kolom Dinamis (Data Integrity)
                    spesifikasi_nama_barang = row[20] if len(row) > 20 else None
                    spesifikasi_lainnya = row[21] if len(row) > 21 else None
                    
                    status_hak_sementara = row[22] if len(row) > 22 else None
                    nomor_hak = row[23] if len(row) > 23 else None
                    nama_pemegang_hak = row[24] if len(row) > 24 else None
                    keterangan_sertifikasi_lainnya = row[25] if len(row) > 25 else None

                    # Validasi Keberadaan Master OPD berdasarkan ID Excel
                    try:
                        opd_obj = MasterOPD.objects.get(id=opd_id)
                    except MasterOPD.DoesNotExist:
                        baris_error_logs.append(f"Baris {i}: ID OPD '{opd_id}' tidak terdaftar di sistem.")
                        gagal += 1
                        continue

                    # Eksekusi Penyuntikan Data ke Django ORM
                    try:
                        # 1. Suntik Data ke Tabel Induk (Master Aset)
                        aset_baru = AsetTanah.objects.create(
                            opd=opd_obj, nama_barang=nama_barang, kode_barang=kode_barang, nibar=nibar,
                            nomor_register=nomor_register, status_kepemilikan=status_kepemilikan,
                            luas_m2=luas_m2 if luas_m2 else 0, cara_perolehan=cara_perolehan,
                            tanggal_perolehan=tanggal_perolehan if tanggal_perolehan else None,
                            nilai_aset=nilai_aset if nilai_aset else 0, latitude=latitude if latitude else 0,
                            longitude=longitude if longitude else 0, kecamatan_nama=kecamatan_nama,
                            kelurahan=kel_obj, kelurahan_nama=kelurahan_nama, alamat_lokasi=alamat_lokasi,
                            status_sertifikasi=status_sertifikasi, nomor_sertifikat=nomor_sertifikat,
                            kondisi_pemanfaatan=kondisi_pemanfaatan, status_penggunaan=status_penggunaan,
                            status_verifikasi=status_verifikasi, satuan='m2',
                            spesifikasi_nama_barang=spesifikasi_nama_barang, spesifikasi_lainnya=spesifikasi_lainnya,
                            status_hak_sementara=status_hak_sementara, nomor_hak=nomor_hak,
                            nama_pemegang_hak=nama_pemegang_hak, keterangan_sertifikasi_lainnya=keterangan_sertifikasi_lainnya
                        )
                        
                        # 🔥 2. LOGIKA OTOMATISASI ANAK (Mencegah Data Yatim Piatu di Web & PDF)
                        if status_sertifikasi == 'BERSERTIFIKAT':
                            # 1. Klasifikasikan kondisi fisik dokumen
                            # SINKRONISASI: Ubah 'FOTOKOPI' menjadi 'FOTO_KOPI' agar pas dengan value HTML kamu!
                            keterangan_fisik = 'FOTO_KOPI'
                            if keterangan_sertifikasi_lainnya in ['ASLI_PEMKOT', 'ASLI_NON_PEMKOT']:
                                keterangan_fisik = 'ASLI'
                            
                            SertifikatTanah.objects.create(
                                aset_tanah=aset_baru, # Menghubungkan relasi ForeignKey
                                opd=opd_obj,
                                nomor_sertifikat=nomor_sertifikat,
                                nama_pemegang_hak=nama_pemegang_hak if nama_pemegang_hak else "-",
                                status_hak=status_hak_sementara if status_hak_sementara else "HAK_PAKAI",
                                nomor_hak=nomor_hak if nomor_hak else "-",
                                keterangan=keterangan_fisik, # Tersinkron ke fungsi logika_kondisi_fisik_dokumen
                                luas=float(luas_m2) if luas_m2 else 0,
                                nilai=float(nilai_aset) if nilai_aset else 0,
                                alamat=alamat_lokasi if alamat_lokasi else "-",
                                peruntukan=spesifikasi_nama_barang if spesifikasi_nama_barang else "-",
                                tanggal_pembuatan=tanggal_perolehan if tanggal_perolehan else None
                            )
                        berhasil += 1
                    except Exception as e:
                        baris_error_logs.append(f"Baris {i} ({nama_barang}): {str(e)}")
                        gagal += 1

                # Jika di akhir iterasi ada data yang gagal, paksa gagalkan semua transaksi
                if gagal > 0:
                    raise Exception("Ada baris data yang tidak valid.")

            messages.success(request, f"Berhasil mengimpor {berhasil} data aset baru ke sistem.")
            return redirect('tanah:list_aset_tanah')

        except Exception as e:
            # Satukan log error untuk ditampilkan di feedback bootstrap
            error_gabungan = " <br> ".join(baris_error_logs) if baris_error_logs else str(e)
            messages.error(request, f"<b>Gagal Mengimpor File Excel!</b> Seluruh operasi dibatalkan.<br>{error_gabungan}")
            return redirect('tanah:halaman_import')

    messages.error(request, "Silakan pilih file Excel terlebih dahulu.")
    return redirect('tanah:halaman_import')

# - - - - -
# PDF - detail
# - - - - -

# 1. pdf profile per aset
@login_required(login_url='auth:login')
@role_required(allowed_roles=['ADMIN_BPKAD', 'SUPERADMIN'])
def export_pdf_detail(request, id_aset):
    # Ambil data aset
    aset = get_object_or_404(AsetTanah.objects.prefetch_related('koleksi_foto'), id=id_aset)
    # Tarik data sertifikat (bisa kosong jika belum sertifikat)
    sertifikat = aset.data_sertifikat.first()
    
    # Logo
    logo_path = os.path.join(settings.BASE_DIR, 'frontend', 'static', 'images', 'logo.png')
    template_path = 'aset_tanah/reports/pdf_profil_aset.html'
    
    # 1. data aset di publik
    path_publik = reverse('publik:data_aset_publik') 
    url_verifikasi = request.build_absolute_uri(path_publik) + f"?id={aset.id}"

    # 2. Generate QR Code secara lokal dalam bentuk biner memori
    qr = qrcode.QRCode(version=1, box_size=10, border=2)
    qr.add_data(url_verifikasi)
    qr.make(fit=True)
    qr_img = qr.make_image(fill_color="#002D5E", back_color="white")

    buffer = BytesIO()
    qr_img.save(buffer, format="PNG")
    qr_base64 = base64.b64encode(buffer.getvalue()).decode()

    luas_indonesia = f"{aset.luas_m2:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
    rupiah_indonesia = f"{aset.nilai_aset:,}".replace(",", ".")

    # Hilangkan buntut ,00 pada luas jika angkanya bulat murni (UX Premium)
    if luas_indonesia.endswith(',00'):
        luas_indonesia = luas_indonesia[:-3]
        
        
    # Render ke HTML khusus PDF
    template_path = 'aset_tanah/reports/pdf_profil_aset.html'
    context = {
        'aset': aset, 
        'sertifikat': sertifikat,
        'user': request.user,
        'logo_path': logo_path,
        'koleksi_foto': aset.koleksi_foto.all(),
        'qr_code_base64': f"data:image/png;base64,{qr_base64}", # Dikirim sebagai Data URI aman
        'url_verifikasi': url_verifikasi,
        
        'luas_terformat': luas_indonesia,
        'nilai_terformat': rupiah_indonesia
    }
    
    response = HttpResponse(content_type='application/pdf')
    # Pakai 'attachment' agar terunduh otomatis
    response['Content-Disposition'] = f'attachment; filename="Profil Aset {aset.kode_barang or aset.id}.pdf"'
    
    template = get_template(template_path)
    html = template.render(context)
    
    # Proses konversi HTML ke PDF
    pisa_status = pisa.CreatePDF(html, dest=response)
    if pisa_status.err:
        return HttpResponse(f'Terjadi kesalahan saat membuat PDF: <pre>{html}</pre>')
        
    return response

# 2. halaman export pdf
@login_required(login_url='auth:login')
@role_required(allowed_roles=['ADMIN_BPKAD', 'SUPERADMIN'])
def halaman_export_pdf(request):
    # Ambil daftar OPD untuk dipasang di dropdown pilihan kartu ke-2
    opd_list = MasterOPD.objects.filter(is_active=1).order_by('nama_opd')
    
    return render(request, 'aset_tanah/reports/halaman_export_pdf.html', {
        'opd_list': opd_list
    })
    
# pdf rekap sertif
@login_required(login_url='auth:login')
@role_required(allowed_roles=['ADMIN_BPKAD', 'SUPERADMIN'])
def export_pdf_rekap_sertif(request):
    # A. Tarik seluruh 40 OPD Aktif & Aset Tanah tervalidasi VALID
    semua_opd = MasterOPD.objects.filter(is_active=1).order_by('nama_opd')
    aset_valid = AsetTanah.objects.filter(status_verifikasi='VALID').select_related('opd')
    
    # B. Inisialisasi hitung akumulasi data
    stats_per_opd = {}
    for opd in semua_opd:
        stats_per_opd[opd.id] = {
            'total_aset': 0, 'sertif_total': 0, 'sertif_luas': 0.0, 'sertif_nilai': 0.0,
            'belum_total': 0, 'belum_luas': 0.0, 'belum_nilai': 0.0,
        }
        
    for aset in aset_valid:
        if aset.opd_id in stats_per_opd:
            stats = stats_per_opd[aset.opd_id]
            stats['total_aset'] += 1
            
            if aset.status_sertifikasi == 'BERSERTIFIKAT':
                stats['sertif_total'] += 1
                stats['sertif_luas'] += float(aset.luas_m2) if aset.luas_m2 else 0.0
                stats['sertif_nilai'] += float(aset.nilai_aset) if aset.nilai_aset else 0.0
            elif aset.status_sertifikasi == 'BELUM_BERSERTIFIKAT':
                stats['belum_total'] += 1
                stats['belum_luas'] += float(aset.luas_m2) if aset.luas_m2 else 0.0
                stats['belum_nilai'] += float(aset.nilai_aset) if aset.nilai_aset else 0.0

    # C. Wadah hitung Grand Total bawah tabel
    grand_total = {
        'aset': 0, 's_total': 0, 's_luas': 0, 's_nilai': 0, 'b_total': 0, 'b_luas': 0, 'b_nilai': 0, 'persen': 0.0
    }
    
    daftar_rekap_raw = []
    for opd in semua_opd:
        stats = stats_per_opd[opd.id]
        total_aset = stats['total_aset']
        persen = (stats['sertif_total'] / total_aset * 100) if total_aset > 0 else 0.0
        
        daftar_rekap_raw.append({
            'nama_opd': opd.nama_opd, 'total_aset': total_aset, 'sertif_total': stats['sertif_total'],
            'sertif_luas': stats['sertif_luas'], 'sertif_nilai': stats['sertif_nilai'],
            'belum_total': stats['belum_total'], 'belum_luas': stats['belum_luas'],
            'belum_nilai': stats['belum_nilai'], 'persen': persen
        })
        
        grand_total['aset'] += total_aset
        grand_total['s_total'] += stats['sertif_total']
        grand_total['s_luas'] += stats['sertif_luas']
        grand_total['s_nilai'] += stats['sertif_nilai']
        grand_total['b_total'] += stats['belum_total']
        grand_total['b_luas'] += stats['belum_luas']
        grand_total['b_nilai'] += stats['belum_nilai']
        
    grand_total['persen'] = (grand_total['s_total'] / grand_total['aset'] * 100) if grand_total['aset'] > 0 else 0.0

    # D. ENGINE FORMATTING: Mengubah angka mulus menjadi format titik (.) khas Indonesia
    def format_ke_indonesia(angka, dengan_rp=False, tampilkan_nol=False):
        if not angka or angka == 0:
            return "Rp 0" if (dengan_rp and tampilkan_nol) else ("0" if tampilkan_nol else "-")
        formatted = f"{int(angka):,}".replace(",", ".")
        return f"Rp {formatted}" if dengan_rp else formatted

    daftar_rekap_final = []
    for r in daftar_rekap_raw:
        daftar_rekap_final.append({
            'nama_opd': r['nama_opd'],
            'total_aset': r['total_aset'],
            'sertif_total': r['sertif_total'],
            'sertif_luas': format_ke_indonesia(r['sertif_luas']),
            'sertif_nilai': format_ke_indonesia(r['sertif_nilai'], dengan_rp=True),
            'belum_total': r['belum_total'],
            'belum_luas': format_ke_indonesia(r['belum_luas']),
            'belum_nilai': format_ke_indonesia(r['belum_nilai'], dengan_rp=True),
            'persen': f"{r['persen']:.2f}%"
        })

    grand_total_final = {
        'aset': format_ke_indonesia(grand_total['aset'], tampilkan_nol=True),
        's_total': format_ke_indonesia(grand_total['s_total'], tampilkan_nol=True),
        's_luas': format_ke_indonesia(grand_total['s_luas'], tampilkan_nol=True),
        's_nilai': format_ke_indonesia(grand_total['s_nilai'], dengan_rp=True, tampilkan_nol=True),
        'b_total': format_ke_indonesia(grand_total['b_total'], tampilkan_nol=True),
        'b_luas': format_ke_indonesia(grand_total['b_luas'], tampilkan_nol=True),
        'b_nilai': format_ke_indonesia(grand_total['b_nilai'], dengan_rp=True, tampilkan_nol=True),
        'persen': f"{grand_total['persen']:.2f}%"
    }

    # E. Bangun Tautan Verifikasi Absolut untuk Sisi Kiri Footer
    url_verifikasi = request.build_absolute_uri(reverse('tanah:halaman_export_pdf'))

    # F. Ambil parameter gambar & Render
    logo_path = os.path.join(settings.BASE_DIR, 'frontend', 'static', 'images', 'logo.png')
    template_path = 'aset_tanah/reports/pdf_rekap_sertif.html'
    
    context = {
        'daftar_rekap': daftar_rekap_final,
        'grand_total': grand_total_final,
        'logo_path': logo_path,
        'url_verifikasi': url_verifikasi,
        'user': request.user
    }
    
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="Rekapitulasi Sertifikasi Aset Tanah.pdf"'
    
    template = get_template(template_path)
    html = template.render(context)
    
    pisa_status = pisa.CreatePDF(html, dest=response)
    if pisa_status.err:
        return HttpResponse(f'Gagal kompilasi PDF: <pre>{html}</pre>')
        
    return response

# pdf rekap opd
@login_required(login_url='auth:login')
@role_required(allowed_roles=['ADMIN_BPKAD', 'SUPERADMIN'])
def export_pdf_aset_opd(request):
    opd_id = request.GET.get('opd_id')
    
    if opd_id:
        opd_target = get_object_or_404(MasterOPD, id=opd_id, is_active=1)
    else:
        opd_target = MasterOPD.objects.filter(is_active=1).first()

    if not opd_target:
        return HttpResponse("Belum ada data OPD yang aktif di sistem.")

    # A. QUERY DINAMIS: Cari akun pengguna yang terhubung ke opd_id ini
    # Kita filter User melalui relasi profilnya (misal nama relation-nya 'profile')
    user_opd = User.objects.filter(profile__opd=opd_target).select_related('profile').first()
    
    if user_opd:
        # Ambil first_name dan last_name dari auth_user
        nama_kepala = f"{user_opd.first_name} {user_opd.last_name}".strip()
        # Ambil NIP dari tabel user_profile
        nip_kepala = user_opd.profile.nip or "-"
    else:
        # Fallback aman jika belum ada user yang didelegasikan ke OPD tersebut
        nama_kepala = getattr(opd_target, 'nama_kepala', "-") or "-"
        nip_kepala = "-"

    # B. Tarik seluruh aset tervalidasi milik OPD terpilih
    aset_list = AsetTanah.objects.filter(opd=opd_target, status_verifikasi='VALID').prefetch_related('data_sertifikat')

    # C. Fungsi Pembantu Format Ribuan Indonesia
    def format_ke_indonesia(angka, dengan_rp=False):
        if not angka or angka == 0:
            return "Rp 0" if dengan_rp else "-"
        formatted = f"{int(angka):,}".replace(",", ".")
        return f"Rp {formatted}" if dengan_rp else formatted
    
    # Memotong string raksasa tanpa spasi agar otomatis wrap down ke bawah
    def potong_string_beruntun(teks, batas=11):
        if not teks or teks == "-":
            return "-"
        teks = str(teks)
        return " ".join(teks[i:i+batas] for i in range(0, len(teks), batas))
    
    # D. Olah Kompilasi Data Baris
    daftar_aset_final = []
    for aset in aset_list:
        sertif = aset.data_sertifikat.first()
        kl = aset.keterangan_sertifikasi_lainnya or ''
        
        status_dokumen = "-"
        status_5_pilar = "Lainnya"
        
        if aset.status_sertifikasi == 'BERSERTIFIKAT':
            ket_fisik = sertif.keterangan if sertif else ''
            if 'ASLI' in kl or 'ASLI' in ket_fisik:
                status_dokumen = "Asli"
            elif 'FOTOKOPI' in kl or 'FC' in kl or (ket_fisik and 'FOTO' in ket_fisik):
                status_dokumen = "FC"
            else:
                status_dokumen = "Asli" if (sertif and sertif.keterangan == 'ASLI') else "FC"

            if 'PEMKOT' in kl or (sertif and 'Pemerintah Kota Palu' in sertif.nama_pemegang_hak):
                status_5_pilar = "Sertifikat Asli Pemkot" if status_dokumen == "Asli" else "Sertifikat FC Pemkot"
            else:
                status_5_pilar = "Sertifikat Asli Non-Pemkot" if status_dokumen == "Asli" else "Sertifikat FC Non-Pemkot"

        lokasi_gabung = f"{aset.alamat_lokasi or ''}, Kec. {aset.kecamatan_nama or '-'}, Kel. {aset.kelurahan_nama or '-'}"
        
        daftar_aset_final.append({
            # Bungkus teks-teks berpotensi raksasa ke dalam engine pemotong string
            'kode_barang': potong_string_beruntun(aset.kode_barang),
            'nama_barang': aset.nama_barang or "-",
            'nibar': potong_string_beruntun(aset.nibar, batas=13),
            'nomor_register': potong_string_beruntun(aset.nomor_register, batas=5),
            'spesifikasi_nama_barang': aset.spesifikasi_nama_barang or "-",
            'spesifikasi_lainnya': aset.spesifikasi_lainnya or "-",
            'luas': format_ke_indonesia(aset.luas_m2),
            'lokasi': lokasi_gabung,
            'koordinat': potong_string_beruntun(f"{aset.latitude or '-'}, {aset.longitude or '-'}", batas=10),
            
            # Sub-Kolom Bukti Registrasi Dokumen
            'sertif_nama': sertif.nama_pemegang_hak if sertif else (aset.nama_pemegang_hak if aset.nama_pemegang_hak else "-"),
            'sertif_nomor': potong_string_beruntun(sertif.nomor_sertifikat if sertif else (aset.nomor_sertifikat if aset.nomor_sertifikat else "-"), batas=11),
            'sertif_tanggal': sertif.tanggal_pembuatan.strftime("%d-%m-%Y") if (sertif and sertif.tanggal_pembuatan) else "-",
            'status_dokumen': status_dokumen,
            'status_5_pilar': status_5_pilar,
            
            # Ekor Baris
            'cara_perolehan': aset.cara_perolehan or "-",
            'status_penggunaan': aset.status_penggunaan or "-",
            'keterangan': aset.keterangan or "-",
            'nilai': format_ke_indonesia(aset.nilai_aset, dengan_rp=True)
        })

    url_verifikasi = request.build_absolute_uri(reverse('tanah:halaman_export_pdf'))
    logo_path = os.path.join(settings.BASE_DIR, 'frontend', 'static', 'images', 'logo.png')
    template_path = 'aset_tanah/reports/pdf_aset_opd.html'

    context = {
        'opd': opd_target,
        'daftar_aset': daftar_aset_final,
        'logo_path': logo_path,
        'url_verifikasi': url_verifikasi,
        'nama_kepala': nama_kepala,
        'nip_kepala': nip_kepala,
        'user': request.user
    }

    # 1. Bersihkan nama OPD dari spasi dan karakter ilegal untuk nama file browser
    nama_opd_clean = opd_target.nama_opd.replace(" ", "_").replace("/", "-").replace("&", "Dan")

    # 2. Setel ke Content-Disposition
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="Buku_Inventaris_Tanah_{nama_opd_clean}.pdf"'
    
    template = get_template(template_path)
    html = template.render(context)
    
    pisa_status = pisa.CreatePDF(html, dest=response)
    if pisa_status.err:
        return HttpResponse(f'Gagal kompilasi PDF Instansi: <pre>{html}</pre>')
        
    return response