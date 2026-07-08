#!/usr/bin/env python
import os
import sys
import django
import openpyxl
from django.db import transaction

# 1. BOOTSTRAP CORE ENGINE DJANGO
# Menghidupkan jantung Django agar terminal bisa mengakses Model & Database
os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'pataba_core.settings')
django.setup()

# Import model-model setelah Django siap
from apps.aset_tanah.models import MasterOPD, AsetTanah, SertifikatTanah, RefKelurahan
from pataba_core.utils import bersihkan_angka # Memanggil sanitizer global kita

def jalankan_migrasi_massal(nama_file_excel):
    if not os.path.exists(nama_file_excel):
        print(f"\n❌ Error: Berkas '{nama_file_excel}' tidak ditemukan di folder utama!")
        return

    print("\n" + "="*60)
    print("🚀 PATABA MASS MIGRATION CLI ENGINE - BPKAD KOTA PALU")
    print("="*60)
    print(f"[*] Membuka file spreadsheet: {nama_file_excel}...")
    
    try:
        wb = openpyxl.load_workbook(nama_file_excel, data_only=True)
        sheet = wb.active
        
        # Hitung estimasi baris untuk progress bar
        total_baris = 0
        for r in sheet.iter_rows(min_row=3, values_only=True):
            if r and r[0] is not None:
                total_baris += 1
            else:
                break
                
        print(f"[*] Terdeteksi total: {total_baris} baris data siap dimigrasikan.")
        print("[*] Memulai transaksi database aman...")
        print("-"*60)

        berhasil = 0
        gagal = 0

        # Gunakan transaction.atomic agar jika di baris 500 error, database tidak kotor
        with transaction.atomic():
            for i, row in enumerate(sheet.iter_rows(min_row=3, values_only=True), start=3):
                if not row or row[0] is None:
                    break

                opd_id = row[0]
                nama_barang = row[1]
                kode_barang = row[2]
                nibar = row[3]
                
                # Standarisasi nomor register otomatis
                nomor_register_raw = row[4]
                if nomor_register_raw is not None:
                    if isinstance(nomor_register_raw, (int, float)):
                        nomor_register = str(int(nomor_register_raw)).strip().zfill(4)
                    else:
                        nomor_register = str(nomor_register_raw).strip().zfill(4)
                else:
                    nomor_register = "0000"

                status_kepemilikan = row[5]
                
                # Jinakkan angka desimal dan rupiah lewat sanitizer global
                luas_m2 = bersihkan_angka(row[6], ke_integer=False)
                cara_perolehan = row[7]
                tanggal_perolehan = row[8]
                nilai_aset = bersihkan_angka(row[9], ke_integer=True)
                
                latitude = row[10]
                longitude = row[11]
                kecamatan_nama = row[12]
                kelurahan_nama = row[13]
                alamat_lokasi = row[14]
                
                kel_obj = None
                if kelurahan_nama:
                    kel_obj = RefKelurahan.objects.filter(nama_kelurahan__iexact=kelurahan_nama).first()
                
                status_sertifikasi = row[15] if row[15] else 'BELUM_BERSERTIFIKAT'
                nomor_sertifikat = row[16]
                kondisi_pemanfaatan = row[17]
                status_penggunaan = row[18]
                status_verifikasi = row[19] if row[19] else 'VALID'

                spesifikasi_nama_barang = row[20] if len(row) > 20 else None
                spesifikasi_lainnya = row[21] if len(row) > 21 else None
                status_hak_sementara = row[22] if len(row) > 22 else None
                nomor_hak = row[23] if len(row) > 23 else None
                nama_pemegang_hak = row[24] if len(row) > 24 else None
                keterangan_sertifikasi_lainnya = row[25] if len(row) > 25 else None

                # Cari Kaitan OPD
                try:
                    opd_obj = MasterOPD.objects.get(id=int(opd_id))
                except MasterOPD.DoesNotExist:
                    print(f"❌ Baris {i}: Gagal! ID OPD '{opd_id}' tidak terdaftar di sistem.")
                    gagal += 1
                    raise Exception(f"ID OPD {opd_id} Tidak Valid.")

                try:
                    # 1. Suntik Data ke Tabel Induk (Master Aset)
                    aset_baru = AsetTanah.objects.create(
                        opd=opd_obj, nama_barang=nama_barang, kode_barang=kode_barang, nibar=nibar,
                        nomor_register=nomor_register, status_kepemilikan=status_kepemilikan,
                        luas_m2=luas_m2, cara_perolehan=cara_perolehan,
                        tanggal_perolehan=tanggal_perolehan if tanggal_perolehan else None,
                        nilai_aset=nilai_aset, latitude=latitude if latitude else 0,
                        longitude=longitude if longitude else 0, kecamatan_nama=kecamatan_nama,
                        kelurahan=kel_obj, kelurahan_nama=kelurahan_nama, alamat_lokasi=alamat_lokasi,
                        status_sertifikasi=status_sertifikasi, nomor_sertifikat=nomor_sertifikat,
                        kondisi_pemanfaatan=kondisi_pemanfaatan, status_penggunaan=status_penggunaan,
                        status_verifikasi=status_verifikasi, satuan='m2',
                        spesifikasi_nama_barang=spesifikasi_nama_barang, spesifikasi_lainnya=spesifikasi_lainnya,
                        status_hak_sementara=status_hak_sementara, nomor_hak=nomor_hak,
                        nama_pemegang_hak=nama_pemegang_hak, 
                        keterangan_sertifikasi_lainnya=keterangan_sertifikasi_lainnya
                    )

                    # 2. Suntik Otomatis Data ke Tabel Anak (SertifikatTanah)
                    if status_sertifikasi == 'BERSERTIFIKAT':
                        keterangan_fisik = 'FOTO_KOPI'
                        if keterangan_sertifikasi_lainnya in ['ASLI_PEMKOT', 'ASLI_NON_PEMKOT']:
                            keterangan_fisik = 'ASLI'

                        SertifikatTanah.objects.create(
                            aset_tanah=aset_baru,
                            opd=opd_obj, # Mengikat OPD dengan aman[cite: 2]
                            nomor_sertifikat=nomor_sertifikat,
                            nama_pemegang_hak=nama_pemegang_hak if nama_pemegang_hak else "Pemerintah Kota Palu",
                            status_hak=status_hak_sementara if status_hak_sementara else "HAK_PAKAI",
                            nomor_hak=nomor_hak if nomor_hak else "-",
                            keterangan=keterangan_fisik, # Terkunci string 'ASLI' / 'FOTO_KOPI'[cite: 2]
                            luas=float(luas_m2) if luas_m2 else 0,
                            nilai=float(nilai_aset) if nilai_aset else 0,
                            alamat=alamat_lokasi if alamat_lokasi else "-",
                            peruntukan=spesifikasi_nama_barang if spesifikasi_nama_barang else "-",
                            tanggal_pembuatan=tanggal_perolehan if tanggal_perolehan else None
                        )

                    berhasil += 1
                    
                    # Cetak live progress counter di terminal agar asyik dipantau
                    sys.stdout.write(f"\r⏳ Progress: [{berhasil}/{total_baris}] Memigrasikan: {nama_barang[:30]}...")
                    sys.stdout.flush()

                except Exception as e:
                    print(f"\n❌ Gagal di Baris {i} ({nama_barang}): {str(e)}")
                    gagal += 1
                    raise e

        print("\n" + "="*60)
        print(f"🎉 MIGRASI SELESAI SECARA SEMPURNA!")
        print(f"   - Berhasil Di-import : {berhasil} aset tanah (+ dokumen anak)")
        print(f"   - Gagal/Dilewati     : {gagal} baris")
        print("="*60 + "\n")

    except Exception as e:
        print(f"\n💥 Transaksi Dibatalkan Global: Ada kesalahan sistem! \nDetail: {str(e)}\n")

if __name__ == '__main__':
    # 📝 UBAH NAMA FILE INI SESUAI DENGAN NAMA BERKAS EXCEL DATASET TIM KAMU
    NAMA_FILE_DATASET = 'dummy2.xlsx' 
    jalankan_migrasi_massal(NAMA_FILE_DATASET)